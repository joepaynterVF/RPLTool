import dash_bootstrap_components as dbc
import dash_uploader as du
import flask
from bs4 import BeautifulSoup
import csv
from dash import Dash, html, dcc, Input, Output, State, callback_context
import plotly.express as px
import pandas as pd
import numpy as np
import os
import openpyxl
import glob
import signal
from sklearn.metrics import euclidean_distances
import dash_daq as daq
from pathlib import Path
import simplekml
from concurrent.futures import ThreadPoolExecutor
from zipfile import ZipFile

# Global Variables
check = False
header_check_cable_all = False
sorted_coordinates = []
unsorted_coordinates = []
graph_name = ""
segment_name = ""
number = 0
cable_end_check = True

# Set colours so that first plot is shown as unsorted, in orange.
initialFig = px.scatter_geo(color_discrete_sequence=["#ef553c", "#636efa"])
initialFig.update_geos(
    resolution=50,
    showframe=True,
    showland=True, landcolor="Green",
    showocean=True, oceancolor="LightBlue",
    showcountries=True, countrycolor="Yellow",
    showlakes=True, lakecolor="LightBlue")
initialFig.update_layout(clickmode='event+select', title=graph_name, title_x=0.55, title_font_family='sans-serif')
suggestedCoordinates = ''
latSuggestion = ''
longSuggestion = ''

clipboardStyle = {"fontSize": 0, 'display': 'inline-block'}
styles = {
    'pre': {
        'border': 'thin lightgrey solid',
        'overflowX': 'scroll'}
}
# Delete sorted and unsorted csv's if they exist.
if os.path.exists('sorted.csv'):
    os.remove("sorted.csv")
if os.path.exists('unsorted.csv'):
    os.remove("unsorted.csv")

# Initialise Server
server = flask.Flask(__name__)
# Build Components
app = Dash(__name__, title="Coordinate Sorter", server=server, suppress_callback_exceptions=True,
           external_stylesheets=[dbc.themes.BOOTSTRAP])

# Get current operating directory
dir_path = os.getcwd()
# Create folder path for upload folder
UPLOAD_FOLDER = os.path.join(dir_path, "KML Files")

# Create upload folder for KML files to be stored
try:
    os.mkdir(UPLOAD_FOLDER)
except FileExistsError:
    pass


def get_kml_file():
    """
    This function retrieves the KML files inside the upload folder.
    Returns: A list of kml filenames
    """
    # Find all files ending in .kml, saving their path
    kml_file = glob.glob(UPLOAD_FOLDER + '/*.kml')
    kml_files = []
    # Save filename of the kml files found, into kml_files list
    for file in kml_file:
        # Append just the filename and not path
        kml_files.append(os.path.basename(file))
    return kml_files


# Configure Dash upload component
du.configure_upload(app, UPLOAD_FOLDER, use_upload_id=False)


def RPL_layout():
    # Dash Layout Customisation
    layout = html.Div(children=[
        html.Div(
            # Page title
            [html.H2(children='Coordinate Sorter', style={'display': 'flex'}),

             # Shutdown Server button
             html.Div(html.Button('Shutdown Server', id='Quit_btn', n_clicks=0,

                                  style={'position': 'absolute',
                                         'right': '0',
                                         'bottom': '25%',
                                         'top': '25%',
                                         'background-color': 'red'})),
             html.Div(id='Output_div'),

             ], style={
                'display': 'flex',
                'flex-direction': 'row',
                'justify-content': 'center',
                'align-items': 'center',
                'align-content': 'center',
                'position': 'relative',

            }),

        html.Div([

            # Update File List Button
            html.Div([
                html.Button('Update File List', id='update_dropdown_button', n_clicks=0)],
                style={
                    'margin': '10px'
                }),

            # File List Dropdown
            html.Div([
                dcc.Dropdown(get_kml_file(), id='dropdown', placeholder='Select KML File', maxHeight=200,
                             optionHeight=50)],
                style={
                    'width': '15em',
                    'margin': '10px',
                    'text-align': 'left',
                }),

            # Add to map Button
            html.Div([
                html.Button(
                    'Add to Map',
                    id='kml-csv-btn-click',
                    n_clicks=0
                ),

                daq.ToggleSwitch(
                id='toggle-switch-cable-check',
                value=True,
                color="Blue",
                label={"label": "Disable/Enable Cable End Checking",
                       "style": {'margin': '10px', 'font-weight': 600, 'font-size': '13px', }},
                labelPosition="right", style={'margin': '10px', 'text-align': 'center', 'position': 'absolute',
                                     'right': '0'})

                ], className='text'),


            html.Div(id='toggle-switch-cable-check-output'),



        ],
            style={
                'display': 'flex',
                'align-items': 'center',
                'justify-content': 'center'
            }),

        # Upload Files Box
        html.Div([
            du.Upload(
                text='Drag and Drop .kmz or .kml files here',
                text_completed='Completed: ',
                pause_button=False,
                cancel_button=True,
                max_file_size=1800,  # 1800 Mb
                filetypes=['kml', 'kmz'],
                id='upload-files-div', default_style={'minHeight': 2, 'lineHeight': 2},
            ),
        ],
            style={
                'textAlign': 'center',
                'width': '30%',
                'display': 'inline-block',
            },
        ),

        html.Div(className='row', children=[

            # Div to center Latitude, Longitude, Auto Create rpl, Start Sorting Coordinates, Toggle Button
            html.Div([
                html.Div(
                    id='container-button-kml-csv2',
                    children=' ',
                    className='text'
                ),
            ], className='four columns'),

            #  Starting long and lat
            html.Div([
                # Input lat long
                dcc.Input(
                    id="lat",
                    type="text",
                    placeholder="Latitude",
                    debounce=True, style={'margin': '10px'}),
                dcc.Input(
                    id="lon",
                    type="text",
                    placeholder="Longitude",
                    debounce=True, style={'margin': '10px'}
                ),

                html.Div([

                    # Auto Create Button
                    html.Button(
                        'Auto Create RPL',
                        id='auto-try-btn',
                        n_clicks=0, style={'margin': '10px'}),

                    # Start Sorting Coordinates Button
                    html.Button(
                        'Start Sorting Coordinates',
                        id='sort-coordinates-btn',
                        n_clicks=0, style={'margin': '10px'}),

                    # Enable/Disable stopping at overlap toggle Button
                    daq.ToggleSwitch(
                        id='my-toggle-switch',
                        value=False,
                        color="Blue",
                        label={"label": "Disable/Enable stopping at overlap",
                               "style": {'margin': '10px', 'font-weight': 600, 'font-size': '13px', }},
                        labelPosition="right", style={'margin': '10px', 'text-align': 'center'}
                    ),
                    html.Div(id='my-toggle-switch-output'),
                    html.Br()
                ],
                    style={
                        'display': 'flex',
                        'align-items': 'center',
                        'justify-content': 'center'
                    }),

                # Loading animations
                html.Div([
                    # Loading symbol for Update Output function (Everything)
                    dcc.Loading(
                        id="loading-auto-cre",
                        type="default",
                        children=[html.Div(id="loading-auto-create")],
                        style={
                            'display': 'block'
                        }
                    ),
                    # Loading symbol for KMZ extraction wait
                    dcc.Loading(
                        id="loading-1",
                        type="default",
                        children=[html.Div(id="loading-output")],
                        style={
                            'display': 'block'
                        }
                    )
                ], style={
                    'margin': '10px'
                }),
                # Download RPL Button
                html.Div([
                    html.Button("Download RPL", id="btn_RPL"),
                    dcc.Download(id="download_RPL")
                ], style={
                    'display': 'block',
                    'margin-top': '50px'
                }),

                # Select a file to convert div
                html.Div(
                    id='coordinate-results',
                    children=' '),

                # Suggested lat
                html.Div(
                    id='suggested-lat',
                    children='',
                    style={'display': 'inline-block'}
                ),
                # Clipboard lat
                dcc.Clipboard(
                    id="clipboard-lat",
                    target_id="suggested-lat",
                    title="Copy Latitude",
                    style={
                        "fontSize": 0,
                    },
                ),
                # Suggest long
                html.Div(
                    id='suggested-long',
                    children='',
                    style={'display': 'inline-block',
                           'marginLeft': '10px',
                           'marginRight': '10px'
                           },
                    className='text'
                ),
                # Clipboard long
                dcc.Clipboard(
                    id="clipboard-long",
                    target_id="suggested-long",
                    title="Copy Longitude",
                    style={
                        "fontSize": 0,
                    },
                    className='text'
                ),
            ], className='text four columns'),

            # Click for coordinate
            html.Div([
                html.Div(
                    id='click-coordinates-div',
                    children='Click Coordinate To Get Latitude and Longitude',
                    className='text'
                ),
                html.Br(),
                # Display click lat
                html.Div(
                    id='click-data-lat',
                    children='',
                    style={'display': 'inline-block'},
                    className='text'
                ),
                # Click clipboard lat
                dcc.Clipboard(
                    id="click-clipboard-lat",
                    target_id="click-data-lat",
                    title="copy",
                    style={
                        "fontSize": 0,
                    }
                ),
                html.Br(),
                # Display click long
                html.Div(
                    id='click-data-long',
                    children='',
                    className='text',
                ),
                # Click clipboard long
                dcc.Clipboard(
                    id="click-clipboard-long",
                    target_id="click-data-long",
                    title="copy",
                    style={
                        "fontSize": 0,
                    },
                ),
            ], className='text four columns'),
        ]),
        dcc.Graph(
            figure=initialFig,
            id='Graph',
            style={'display': 'inline-block', 'width': '100%', 'height': '50em'},
        ),
    ], style={'text-align': 'center'})
    return layout


@app.callback(
    Output("Output_div", "children"),
    Input("Quit_btn", "n_clicks")
)
def quit1(n_clicks):
    """
    Function to kill the program if shutdown button pressed.
    """
    if n_clicks > 0:
        # Kill program if Quit button pressed
        os.kill(os.getpid(), signal.SIGTERM)
    return

@app.callback(
    Output("download_RPL", "data"),
    Input("btn_RPL", "n_clicks"),
    prevent_initial_call=True,
)
def download_RPL(n_clicks):
    """
    Function to return the RPL to the user as a downloaded file.
    Return: Newly created RPL file.
    """
    global sorted_coordinates
    # Clear coordinates array to avoid added RPL to end of Old RPL if server not restarted
    sorted_coordinates.clear()

    return dcc.send_file(
        'RPL_' + graph_name + '.xls'
    )


def update_options(existing_options):
    """
    This function updates the dropdown box.
    Returns: KML filename
    Params: existing_options: Current list of kml filenames in the dropdown
    """
    # List of kml filenames
    option_name = get_kml_file()
    # Iterate through filenames
    for filename in option_name:
        # Check if filename is in existing_options
        if filename not in existing_options:
            # If filename not in existing_options, append to existing_options
            existing_options.append(filename)
        else:
            # Else just return kml filename
            return option_name

    return option_name

def extract_KML(folder):
    """
    This function extracts kml files from a kmz.
    Returns: Nothing. Saves to upload folder
    Params: folder:Contains the entire contents of a single folder from the kmz
    """

    # Create KML object
    kml = simplekml.Kml()
    # Find name of folder
    folder_name = folder.find('name')
    # Initialise Variables
    placemark_coordinates = []
    system_name = []
    ship_operation = []
    operation_date = []
    cable_type = []
    slack_percent = []
    segment_name = []
    system_type = []
    installation_year = []
    out_of_service_year = []
    allsimpledataobjs = {}

    if folder.find_all('Placemark'):
        # Find all placemarks in folder
        placemarks = folder.find_all('Placemark')
        # For each placemark object in each placemark
        for placemarkobj in placemarks:
            if placemarkobj.find_all('SimpleData'):
                # Find all simpledata tags
                simpledata = placemarkobj.find_all("SimpleData")

                # Cycle through each simpledata tag, saving to allsimpledataobjs dict
                for simpledataobj in simpledata:
                    # Save each simpledata attribute name and the associated value
                    allsimpledataobjs[simpledataobj.get('name')] = simpledataobj.text

                # --------------- Simple data tag gathering --------------- #
                # Cycle through each simpledata attribute adding the value to associate attribute array
                if "system_name" in allsimpledataobjs:
                    systemname = allsimpledataobjs["system_name"]
                    # Clean value so XML does not throw an error.
                    systemname = clean(systemname)
                    system_name.append(systemname)
                else:
                    # If attribute not in simpledata array within KML, append empty field
                    system_name.append("")
                if "ship_operation" in allsimpledataobjs:
                    shipoperation = allsimpledataobjs["ship_operation"]
                    shipoperation = clean(shipoperation)
                    ship_operation.append(shipoperation)
                else:
                    ship_operation.append("")
                if "operation_date" in allsimpledataobjs:
                    operationdate = allsimpledataobjs["operation_date"]
                    operationdate = clean(operationdate)
                    operation_date.append(operationdate)
                else:
                    operation_date.append("")
                if "cable_type" in allsimpledataobjs:
                    cable_type.append(allsimpledataobjs["cable_type"])
                else:
                    cable_type.append("")
                if "slack_percent" in allsimpledataobjs:
                    slackpercent = allsimpledataobjs["slack_percent"]
                    slackpercent = clean(slackpercent)
                    slack_percent.append(slackpercent)
                else:
                    slack_percent.append("")
                if "segment_name" in allsimpledataobjs:
                    segmentname = allsimpledataobjs["segment_name"]
                    segmentname = clean(segmentname)
                    segment_name.append(segmentname)
                else:
                    segment_name.append("")
                if "system_type" in allsimpledataobjs:
                    systemtype = allsimpledataobjs["system_type"]
                    systemtype = clean(systemtype)
                    system_type.append(systemtype)
                else:
                    system_type.append("")
                if "installation_year" in allsimpledataobjs:
                    installationyear = allsimpledataobjs["installation_year"]
                    installationyear = clean(installationyear)
                    installation_year.append(installationyear)
                else:
                    installation_year.append("")
                if "out_of_service_year" in allsimpledataobjs:
                    outofserviceyear = allsimpledataobjs["out_of_service_year"]
                    outofserviceyear = clean(outofserviceyear)
                    out_of_service_year.append(outofserviceyear)
                else:
                    out_of_service_year.append("")
                # Save coordinates within this placemark
                placemark_coordinates.append(placemarkobj.find("coordinates"))
                # Clear allsimpledataobjs, ready for next placemark within folder
                allsimpledataobjs.clear()

        i = 0
        while i <= len(placemark_coordinates):

            try:
                # For each set of coordinates in each placemark, add a new linestring to the kml
                linestring = kml.newlinestring(name=folder_name,coords=process_coordinate_string(placemark_coordinates[i].string, True, "",""))

                # For each set of simpledata attributes in each placemark, add new simpledata field with corrospoding value
                linestring.extendeddata.schemadata.newsimpledata('system_name', system_name[i])
                linestring.extendeddata.schemadata.newsimpledata('ship_operation', ship_operation[i])
                linestring.extendeddata.schemadata.newsimpledata('operation_date', operation_date[i])
                linestring.extendeddata.schemadata.newsimpledata('cable_type', cable_type[i])
                linestring.extendeddata.schemadata.newsimpledata('slack_percent', slack_percent[i])
                linestring.extendeddata.schemadata.newsimpledata('segment_name', segment_name[i])
                linestring.extendeddata.schemadata.newsimpledata('system_type', system_type[i])
                linestring.extendeddata.schemadata.newsimpledata('installation_year', installation_year[i])
                linestring.extendeddata.schemadata.newsimpledata('out_of_service_year', out_of_service_year[i])
            except IndexError:
                break
            # Increment i to cycle through placemark_coordinates and all the arrays
            i += 1

    # Remove chars from folder name to avoid error when creating a path and saving file
    folder_name.string = folder_name.string.replace('/', '')
    folder_name.string = folder_name.string.replace('\\', '')
    folder_name.string = folder_name.string.replace('&', '')
    folder_name.string = folder_name.string.replace('<', '')
    folder_name.string = folder_name.string.replace('>', '')
    folder_name.string = folder_name.string.replace('"', '')
    folder_name.string = folder_name.string.replace("'", '')

    # Set KML name and save directory
    kmlname = os.path.join(UPLOAD_FOLDER, (folder_name.string + ".kml"))
    # Save KML
    kml.save(kmlname)

def clean(var):
    """
    This function cleans the variable of /, \\, &, <, >, ", '
    Params: var: Variable being passed in
    Returns var, clean variable being returned.
    """
    var = var.replace('/', '')
    var = var.replace('\\', '')
    var = var.replace('&', '')
    var = var.replace('<', '')
    var = var.replace('>', '')
    var = var.replace('"', '')
    var = var.replace("'", '')
    return var


@app.callback(Output('loading-output', 'children'),
              [Input('upload-files-div', 'isCompleted')],
              [State('upload-files-div', 'fileNames')], )
def upload_files(isCompleted, fileNames):
    """
    This function uploads and detects kmz files.
    Params: isCompleted: True or False from Dash Upload component.
            fileNames: List of filenames uploaded from Dash Upload component
    Returns: Empty String. Responsible for starting extraction of KMZ file.

    """
    if not isCompleted:
        # Returning list of length 2 to avoid loading error
        return ["", ""]
    if fileNames is not None:
        out = []
        # For each filename in fileNames check if .kmz.
        for filename in fileNames:
            if os.path.splitext(filename)[1] == ".kmz":
                # Unzip kmz to reveal doc.kml
                kmz = ZipFile(os.path.join(UPLOAD_FOLDER, filename), 'r')
                # Open doc.kml
                kml_filename = kmz.open('doc.kml', 'r')
                # Parse kml file
                s = BeautifulSoup(kml_filename, 'xml')
                # Find all folders
                folder_list = s.find_all('Folder')
                # Execute extract_KML() function for each folder in a different thread, parallel tasking, improving performance.
                with ThreadPoolExecutor() as executor:
                    [executor.submit(extract_KML(folder=folder)) for folder in folder_list]
            else:
                # Save file
                file = Path(UPLOAD_FOLDER) / filename
                out.append(file)
        return ["", ""]
    return [html.Ul(html.Li("No Files Uploaded Yet!"))]


@app.callback(
    Output('click-data-long', 'children'),
    Output('click-data-lat', 'children'),
    Output('click-clipboard-long', 'style'),
    Output('click-clipboard-lat', 'style'),
    Input('Graph', 'clickData'))
def display_click_data(clickData):
    """
    This function returns the longitude and latitude of the point the user clicks on the map.
    Params: clickData: Longitude and Latitude of point clicked.
    Returns: Longitude and Latitude of the user selected point.
    """
    if clickData is None:
        return ' ', '', {"fontSize": 0, 'display': 'inline-block'}, {"fontSize": 0, 'display': 'inline-block'}
    else:
        return (
            clickData["points"][0]["lon"], clickData["points"][0]["lat"], {"fontSize": 20, 'display': 'inline-block'},
            {"fontSize": 20, 'display': 'inline-block'})

@app.callback(
    Output('toggle-switch-cable-check-output', 'children'),
    Output('my-toggle-switch-output', 'children'),
    Output('btn_RPL', 'style'),
    Output('dropdown', 'options'),
    Output('coordinate-results', 'children'),
    Output('clipboard-lat', 'style'),
    Output('suggested-lat', 'children'),
    Output('clipboard-long', 'style'),
    Output('suggested-long', 'children'),
    Output('Graph', 'figure'),
    Output('sort-coordinates-btn', 'disabled'),
    Output('auto-try-btn', 'disabled'),
    Output('loading-auto-create', 'children'),
    Input('dropdown', 'value'),
    Input('kml-csv-btn-click', 'n_clicks'),
    Input('sort-coordinates-btn', 'n_clicks'),
    Input("lat", "value"),
    Input("lon", "value"),
    [State('dropdown', 'options')],
    Input('update_dropdown_button', 'n_clicks'),
    Input('auto-try-btn', 'n_clicks'),
    Input('my-toggle-switch', 'value'),
    Input('toggle-switch-cable-check', 'value')

)
def update_output(value, KML_clicks, sort_clicks, lat, lon, existing_options, dropdown_clicks, auto_try_clicks, toggle, cabletoggle):
    """
    This function controls almost every output onto the Dash display. Any click the user makes, flows through this function.
    Params: value: Selected kml from dropdown. Contains filename of kml. If not selected = None.
            KML_clicks: Number of times the Add to Map button has been pressed.
            sort_clicks: Number of times the Start Sorting Coordinates button has been preseed.
            lat: Contains the current Latitude in the latitude input box.
            lon: Contains the current Longitude in the Longitude input box.
            existing_options: Contains the list of options currently in the dropdown.
            dropdown_clicks: Number of times the Update File List button has been pressed.
            auto_try_clicks: Number of times Auto Create RPL button has been pressed.
            toggle: True or False for the Disable/Enable stopping at overlap toggle. Default = False.
    Returns: All of the above parameters with the corresponding value depending on the action that called the function.
    """
    global initialFig
    global graph_name
    global suggestedCoordinates
    global latSuggestion
    global longSuggestion
    global clipboardStyle
    global sorted_coordinates
    global unsorted_coordinates
    global check
    global cable_end_check
    # Find which action has been changed/pressed.
    changed_id = [p['prop_id'] for p in callback_context.triggered][0]

    # Initialise the current kml files.
    kmlfiles = update_options(existing_options)

    # Check if Add to Map button has been pressed or Disable/Enable stopping at overlap has changed
    if 'kml-csv-btn-click' in changed_id or 'my-toggle-switch' in changed_id:
        if value is None:
            # Send "Select a file to convert" message to the user
            return cabletoggle, toggle, {
                'display': 'none'}, kmlfiles, 'Select a file to convert', clipboardStyle, longSuggestion, clipboardStyle, latSuggestion, initialFig, False, False, ""
        # Send selected kml filename to convertKMLToCSV function. Creates a csv containing all the coordinates of kml file
        convertKMLToCSV(value)
        # Set column names for dataframe
        colnames = ['latitude', 'longitude', 'cable_type', 'slack']
        # Check if file exists
        if not os.path.exists('coordinates.csv'):
            return cabletoggle, toggle, {
                'display': 'none'}, kmlfiles, 'An error occurred', clipboardStyle, longSuggestion, clipboardStyle, latSuggestion, initialFig, False, False, ""
        # Save coordinates into a dataframe using the csv convertKMLToCSV function created
        df = pd.read_csv('coordinates.csv', names=colnames, header=None)
        # Initialise map with the dataframe of coordinates
        fig = px.scatter_geo(df, lat='latitude', lon='longitude', color_discrete_sequence=["#ef553c", "#636efa"])
        # Set graph characteristics, like title.
        fig.update_layout(title=graph_name, title_x=0.55, title_font_family='sans-serif')
        # Set size of marker
        fig.update_traces(marker=dict(size=4))
        # Center map view to the cable that has been plotted
        fig.update_geos(fitbounds="locations",
                        resolution=50,
                        showframe=True,
                        showland=True, landcolor="Green",
                        showocean=True, oceancolor="LightBlue",
                        showcountries=True, countrycolor="Yellow",
                        showlakes=True, lakecolor="LightBlue")
        # Apply fig to the global figure
        initialFig = fig
        return cabletoggle, toggle, {
            'display': 'none'}, kmlfiles, '', {'display': 'none'}, "", {
                   'display': 'none'}, "", fig, False, False, ""
    if 'toggle-switch-cable-check' in changed_id:
        cable_end_check = cabletoggle

    # Check if Sort Coordinates button or Auto Create RPL button has been pressed
    if 'sort-coordinates-btn' in changed_id or 'auto-try-btn' in changed_id:
        check = True
        # Check if Lon and Lat are not None
        if lon and lat:
            # Check lon and lat are numbers
            try:
                int(lon)
                int(lat)
            except ValueError:
                try:
                    float(lon)
                    float(lat)
                except ValueError:
                    return cabletoggle, toggle, {
                        'display': 'none'}, kmlfiles, 'Please ensure the coordinates are numbers', clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, initialFig, False, False, ""
            # Save order_section result
            result = order_section(lat, lon)
            # Save coordinates to be used in next loop
            prev_coor = [(lon, lat)]
            # Check if what position toggle button is in
            if cable_end_check is True:
                # If a coordinate has been selected in the middle and auto try button pressed
                startormiddle = findNextPoint(result[0],result[1], "StartOrEnd")
                if type(startormiddle) == str:
                    if startormiddle == "Not an end of cable":
                        return cabletoggle, toggle, {
                                'display': 'none'}, kmlfiles, 'Please select a coordinate at the start or end of the cable', clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, initialFig, False, False, ""
            # Save result to be used in next loop
            prev_result = result
            l = 0
            # If there is an overlap, result returns a list
            while type(result) == list:
                # Create unsorted dataframe from unsorted.csv
                unsorteddf = pd.read_csv(f'unsorted.csv')
                # Create sorted dataframe from sorted.csv
                sorteddf = pd.read_csv(f'sorted.csv')
                # Combine sorted and unsorted dataframes
                df_res = pd.concat([sorteddf, unsorteddf])
                # Plot both sorted and unsorted dataframes
                fig = px.scatter_geo(df_res, lat='Latitude', lon='Longitude', hover_name="Colour", color="Colour",
                                     color_discrete_sequence=["#636efa", "#ef553c"], labels={"Colour": "Key"})
                fig.update_layout(title=graph_name, title_x=0.55, title_font_family='sans-serif')
                fig.update_geos(fitbounds="locations",
                                resolution=50,
                                showframe=True,
                                showland=True, landcolor="Green",
                                showocean=True, oceancolor="LightBlue",
                                showcountries=True, countrycolor="Yellow",
                                showlakes=True, lakecolor="LightBlue")
                fig.update_traces(marker=dict(size=4))
                initialFig = fig
                # Check lon and lat are numbers
                try:
                    int(result[0])
                    int(result[1])
                except ValueError:
                    try:
                        float(result[0])
                        float(result[1])
                    except ValueError:
                        return cabletoggle, toggle, {
                            'display': 'none'}, kmlfiles, 'Please ensure the coordinates are numbers', {'display': 'none'}, "", {'display': 'none'}, "", initialFig, False, False, ""
                # Get suggested next coordinates by sending lat and lon
                suggestedNextCoordinates = findNextPoint(result[0], result[1], "")
                # The style of the coordinates to be copied by the user
                clipboardStyle = {"fontSize": 20, 'display': 'inline-block'}
                # Save suggested next coordinates individually as lat and long
                latSuggestion = suggestedNextCoordinates[0]
                longSuggestion = suggestedNextCoordinates[1]
                # Create response message for suggested coordinates to be output to the user.
                suggestedCoordinates = (
                    'Found gap, last known coordinates: {} , {} '.format(result[0], result[1]), html.Br(),
                    ' Suggest coordinates:')
                # Add coordinates to previous coordinates list
                prev_coor.append((result[0], result[1]))
                # Get new result based on the suggestion of next coordinates
                result = order_section(latSuggestion, longSuggestion)

                # Cycle through previous coordinates until no longer an overlap if auto try button pressed
                if prev_result == result:
                    if 'auto-try-btn' in changed_id:

                        # Decrease counter to count back through previous coordinates
                        l -= 1
                        # Attempt to order section with previous coordinates, until no overlap.
                        result = order_section(prev_coor[l][0], prev_coor[l][1])
                        # Attempt to order section with next coordinates if previous coordinates did not work.
                        if prev_coor[len(prev_coor) - 3] == prev_coor[len(prev_coor) - 2] and prev_coor[len(prev_coor) - 2] == prev_coor[len(prev_coor) - 1]:
                            orig_lat_suggested = latSuggestion
                            orig_long_suggested = longSuggestion
                            # Get coordinate after the next suggested.
                            suggestedNextCoordinates = findNextPoint(latSuggestion, longSuggestion, "forward")
                            latSuggestion = suggestedNextCoordinates[0]
                            longSuggestion = suggestedNextCoordinates[1]
                            result = order_section(latSuggestion, longSuggestion)
                            # If result is equal to suggestednextcoordinates, then sort using the previous original
                            # suggested coordinates, but back one coordinate as forward did not work.
                            if suggestedNextCoordinates.tolist() == result:
                                suggestedNextCoordinates = findNextPoint(orig_lat_suggested, orig_long_suggested, "back")
                                latSuggestion = suggestedNextCoordinates[0]
                                longSuggestion = suggestedNextCoordinates[1]
                                result = order_section(latSuggestion, longSuggestion)

                        # Prevent Infinite Loop
                        if prev_coor[l] == prev_coor[l - 1]:
                            # Delete previous entry as duplicate
                            del prev_coor[-1]
                            return cabletoggle, toggle, {
                                'display': 'none'}, kmlfiles, suggestedCoordinates, clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, initialFig, False, False, ""


                if toggle == True:
                    # Indicates an overlap
                    if prev_result == result:
                        # Stop at the overlap and ask user to enter next coordinates
                        return cabletoggle, toggle, {
                            'display': 'none'}, kmlfiles, suggestedCoordinates, clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, fig, False, False, ""
                    # Update previous result
                    prev_result = result
                    # Increment counter
                    l += 1
                elif toggle == False and 'auto-try-btn' not in changed_id:
                    # Update previous result
                    prev_result = result
                    # Increment counter
                    l += 1
                    return cabletoggle, toggle, {
                        'display': 'none'}, kmlfiles, suggestedCoordinates, clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, fig, False, False, ""
                else:
                    # Update previous result
                    prev_result = result
                    # Increment counter
                    l += 1

            else:
                # Save sorted.csv into a dataframe
                sorteddf2 = pd.read_csv(f'sorted.csv')

                # Initialise graph with sorted.csv
                fig = px.scatter_geo(sorteddf2, lat='Latitude', lon='Longitude', hover_name="Colour", color="Colour",
                                     color_discrete_sequence=["#636efa", "#ef553c"], labels={"Colour": "Key"})
                fig.update_layout(title=graph_name, title_x=0.55)
                fig.update_traces(marker=dict(size=4))
                fig.update_geos(fitbounds="locations",
                                resolution=50,
                                showframe=True,
                                showland=True, landcolor="Green",
                                showocean=True, oceancolor="LightBlue",
                                showcountries=True, countrycolor="Yellow",
                                showlakes=True, lakecolor="LightBlue")

                # Save sorted.csv into a dataframe without colour column
                sortedCoordinates = (pd.read_csv(f'sorted.csv')).drop(['Colour'], axis=1)
                # Create RPL with sorted coordinates
                creatRPLTemplate(sortedCoordinates)
                # Convert dataframe into numpy array
                sortedCoordinates = sortedCoordinates.to_numpy()
                # Create a csv holding the sorted coordinates
                with open("sortedCoordinates.csv", "w") as my_csv:
                    csvWriter = csv.writer(my_csv, delimiter=',')
                    for x in sortedCoordinates:
                        csvWriter.writerow(x)
                # Delete sorted.csv
                os.remove("sorted.csv")
                os.remove("sortedCoordinates.csv")
                initialFig = fig
                # Clear sorted coordinates to avoid multiple cable routes being displayed after sorting more than 1 cable without restarting server
                sorted_coordinates.clear()
                return cabletoggle, toggle, {'display': 'inline-block', }, kmlfiles, result, {"fontSize": 0,
                                                                                 'display': 'inline-block'}, '', {
                           "fontSize": 0, 'display': 'inline-block'}, '', fig, True, True, ""
        else:
            return cabletoggle, toggle, {
                'display': 'none'}, kmlfiles, 'Please enter coordinates', clipboardStyle, longSuggestion, clipboardStyle, latSuggestion, initialFig, False, False, ""
    else:
        if check:
            check = False
            return cabletoggle, toggle, {
            'display': 'none'}, kmlfiles, suggestedCoordinates, clipboardStyle, latSuggestion, clipboardStyle, longSuggestion, initialFig, False, False, ""
        else:
            return cabletoggle, toggle, {
                'display': 'none'}, kmlfiles, '', {'display': 'none'}, "", {
                       'display': 'none'}, "", initialFig, False, False, ""



def creatRPLTemplate(coordinates):
    """
    This function creates the final RPL files to be downloaded by the user.
    Params: coordinates: Dataframe of coordinates
    Returns: Nothing. Directly creates RPL as .xls in the directory.
    """
    global graph_name
    global dir_path
    global segment_name
    # Load RPL template
    wrkbk = openpyxl.load_workbook('RPLTemplate.xlsx')
    # Load current sheet in RPLTemplate, there is only one sheet
    sh = wrkbk.active
    # Template data starts from row 7 onwards as the template contains a header
    i = 7
    # Loop through dataframe of coordinates
    for index, row in coordinates.iterrows():
        # Split integer and fractional parts of longitude and latitude using string method as this is the most accurate.
        strlat = str(row['Latitude'])
        strlat = strlat.split(".")

        LatWhole = int(strlat[0])
        latFrac = "." + strlat[1]
        latFrac = float(latFrac)

        strlong = str(row['Longitude'])
        strlong = strlong.split(".")

        longWhole = int(strlong[0])
        longFrac = "." + strlong[1]
        longFrac = float(longFrac)

        # lat
        sh.cell(row=i, column=3).value = LatWhole
        sh.cell(row=i, column=4).value = latFrac * 60
        sh.cell(row=i, column=5).value = 'N' if LatWhole > 0 else 'S'
        # long
        sh.cell(row=i, column=6).value = longWhole
        sh.cell(row=i, column=7).value = longFrac * 60
        sh.cell(row=i, column=8).value = 'E' if longWhole > 0 else 'W'
        # cable type (row=i+1 as cable type is on row below coordinates)
        sh.cell(row=i+1, column=27).value = row['Cable_Type']
        # slack
        sh.cell(row=i+1, column=24).value = row['Slack']
        # Increment row by 2 as next row in RPLTemplate should not be filled
        i += 2
    # Calculate number of rows left
    num_rows_left = sh.max_row - i
    # Delete rows left
    sh.delete_rows(i, (num_rows_left + 1))

    # ------------ Rename sheet name with segment name -------------- #

    if "seg" in graph_name or "Seg" in graph_name:
        try:
            segment_name = graph_name.split("seg", 1)[1]
        except:
            segment_name = graph_name.split("Seg", 1)[1]
    elif segment_name:
        segment_name = segment_name
    else:
        segment_name = "Unknown Segment"

    # Change worksheet title
    sh.title = segment_name

    # ------------ Change references from old worksheet name to new, fixing broken links -------------- #

    # Iterate through all name ranges
    for i in range(len(wrkbk.defined_names.definedName)):
        # Try to access destinations where a worksheet will be referenced, if not pass
        try:
            # If destination of define ranges equals old worksheet name
            if list(wrkbk.defined_names.definedName[i].destinations)[0][0] == 'Segment_Name':
                # Set old name range name
                old_range_name = str(list(wrkbk.defined_names.definedName[i])[0][1])
                # Set old name range attr text
                old_attr_text = list(wrkbk.defined_names.definedName[i].destinations)[0][1]
                # Delete old name range before we create a new one
                del wrkbk.defined_names.definedName[i]
                # Create new attr text for the new name range
                new_attr_text = "'" + segment_name + "'" + "!" + old_attr_text
                # Create the new name range object with old name range name as the name and new attr text
                A_New_Range = openpyxl.workbook.defined_name.DefinedName(old_range_name, attr_text=new_attr_text)
                # Add new name range to workbook with the updated reference to the new sheet name.
                wrkbk.defined_names.append(A_New_Range)

        except IndexError:
            # If destinations were none, pass.
            continue

    # Save workbook
    wrkbk.save('RPL_' + graph_name + '.xlsx')

    # Get filepath of new RPL
    filepath = os.path.join(dir_path, ('RPL_' + graph_name + '.xlsx'))
    # Convert to .xls file
    pre, ext = os.path.splitext(filepath)
    try:
        os.rename(filepath, pre + '.xls')
    except FileExistsError:
        # Delete file if already exists
        os.remove(os.path.join(dir_path, ('RPL_' + graph_name + '.xls')))
        os.rename(filepath, pre + '.xls')


def process_coordinate_string(coor_str, kmz, cable_type, slack):
    """
    This function breaks the string of coordinates into [Lat,Lon,Cable_Type,Slack,Lat,Lon,Cable_Type,Slack...] for a CSV row
    Params: coor_str: String of coordinates
            kmz: If KMZ or not. bool, True or False
            cable_type: Cable type for coordinates
            slack: Slack for coordinates
    Returns: A list of lists containing latitude and longitude coordinates. [[lat, lon,Cable_Type,Slack],[lat, lon,Cable_Type,Slack]]
    """

    # Clean string of \n and \t
    coor_str = coor_str.strip()

    # Split string by spaces
    unclean_space_splits = coor_str.split(" ")
    space_splits = []
    # Loop through unclean_space_splits
    for x in unclean_space_splits:
        # Check if a section is a digit
        if x[1:2].isdigit() or x[0:1].isdigit():
            # If digit add to space_splits
            space_splits.append(x)

    # Set number of rows as length of space_splits list
    rows = len(space_splits)
    ret = [[0 for x in range(4)] for y in range(rows)]
    for count, value in enumerate(space_splits, start=0):
        comma_split = value.split(',')
        if kmz is True:
            if slack != "":
                ret[count][3] = (slack)  # slack
            if cable_type != "":
                ret[count][2] = (cable_type)  # cable_type

            ret[count][1] = (comma_split[1])  # lat
            ret[count][0] = (comma_split[0])  # lng
        else:
            ret[count][0] = (comma_split[1])  # lat
            ret[count][1] = (comma_split[0])  # lng
            if slack != "":
                ret[count][3] = (slack) # slack
            if cable_type != "":
                ret[count][2] = (cable_type)  # cable_type
    return ret


def convertKMLToCSV(value):
    """
    This function converts KML into a CSV of coordinates from the KML. It finds the slack, cable type, coordinates
    and segment name from the KML.
    Params: Value: KML filename
    Returns: Nothing as it creates a coordinates.csv in the file directory
    """
    global graph_name
    global unsorted_coordinates
    global segment_name
    global header_check_cable_all

    header_check_cable_all = False
    # Empty unsorted coordinates variable
    unsorted_coordinates.clear()
    # If .csv exists, delete
    if os.path.exists("coordinates.csv"):
        os.remove("coordinates.csv")
    if os.path.exists("coordinates_Cable_Type.csv"):
        os.remove("coordinates_Cable_Type.csv")
    if os.path.exists("coordinates_no_cable_type.csv"):
        os.remove("coordinates_no_cable_type.csv")
    if os.path.exists("cable_coordinates_all.csv"):
        os.remove("cable_coordinates_all.csv")

    # Create directory for kml filename
    kml_filename = os.path.join(UPLOAD_FOLDER, value)
    allsimpledataobjs = {}

    # ---- Save cable types from cable_types.csv into cable types array ---- #
    # Load all cable types from cable_types.csv into cable_types_list
    with open('cable_types.csv', newline='') as csvfile:
        cable_types_list = list(csv.reader(csvfile))

    cable_types = []
    # Iterate through sublist within cable_types_list
    for sublist in cable_types_list:
        # Iterate through each item within the sublist
        for item in sublist:
            # Append each cable type into cable_types array
            cable_types.append(item)

    # Open kml file
    with open(kml_filename, 'r') as f:
        s = BeautifulSoup(f, 'xml')

        # ---- Save graph name ---- #
        # Find name in kml file
        if s.find('Schema'):
            graph_name = s.find('Schema')['name']
        else:
            # Save graph name as file name if no name in kml
            # Save first index 0 from split as index 1 is the file extension, .kml
            graph_name = os.path.basename(os.path.splitext(kml_filename)[0])

        # Check if there are placemarks in KML
        if s.find_all('Placemark'):
            # Find all placemarks
            placemark = s.find_all('Placemark')
            # Iterate through each placemark
            for placemarkobj in placemark:
                # Check if placemark object contains simpledata fields
                if placemarkobj.find_all('SimpleData'):
                    # Find all simpledata tags
                    simpledata = placemarkobj.find_all("SimpleData")

                    # Cycle through each simpledata tag, saving to allsimpledataobjs dict
                    for simpledataobj in simpledata:
                        # Save each simpledata attribute name and the associated value
                        allsimpledataobjs[simpledataobj.get('name')] = simpledataobj.text

                    # ------ Simple data tag gathering ------ #
                    """
                    Cycle through each simpledata attribute adding the value to associated attribute array
                    """
                    # ---- Get cable type ---- #
                    if "cable_type" in allsimpledataobjs:
                        cable_type = allsimpledataobjs["cable_type"]
                    elif "system_cable_types" in allsimpledataobjs:
                        cable_type = allsimpledataobjs["system_cable_types"]
                    else:
                        cable_type = ""

                    # ---- Get slack ---- #
                    if "slack_percent" in allsimpledataobjs:
                        # Find slack in simple data
                        slackpercent = allsimpledataobjs["slack_percent"]
                        # Clean
                        slackpercent = clean(slackpercent)
                        slack_percent = slackpercent
                    else:
                        slack_percent = ""

                    # ---- Get segment name ---- #
                    if "segment_name" in allsimpledataobjs:
                        # Find segment name in simple data
                        segmentname = allsimpledataobjs["segment_name"]
                        # Clean
                        segmentname = clean(segmentname)
                        segment_name = segmentname
                    else:
                        segment_name = ""

                    # Send coordinates, cable type and slack to CoordsToCSV
                    CoordsToCSV(placemarkobj.find("coordinates"), cable_type, slack_percent)
                    # Clear allsimpledataobjs, ready for next placemark within folder
                    allsimpledataobjs.clear()

                # ---- If cable types are saved as placemark names ---- #
                elif placemarkobj.find("name").text in cable_types:
                    # Get coordinates
                    coordinates = placemarkobj.find_all("coordinates")
                    # Put coordinates into CSV
                    CoordsToCSV(coordinates, placemarkobj.find("name").text, "")

                else:
                    # --------- No cable types found --------- #

                    # If a point, don't save coordinate as it is not part of the RPL
                    if placemarkobj.find_all("Point"):
                        continue
                    # Put coordinates into CSV
                    CoordsToCSV(placemarkobj.find_all("coordinates"), "", "")

        else:
            # --------- No placemarks found --------- #
            CoordsToCSV(s.find_all('coordinates'), "", "")

    CheckCSV()

def CheckCSV():
    """
    This function checks for any duplicates within the coordinates and favours the cable type coordinates over the non
    cable type coordinates.
    Return: Creates coordinates.csv
    """

    # Check if csv is present
    if not os.path.exists("cable_coordinates_all.csv"):
        return "An error occurred"

    # ---------- Attempt number 1 to remove duplicates ---------- #
    # Load csv into dataframe
    first_df = pd.read_csv('cable_coordinates_all.csv')
    # Replace all 0 with NaN
    first_df.replace('0', np.nan, inplace=True)
    # Sort values so that all NaN are first
    first_df = first_df.sort_values('Cable_Type', na_position='first')
    # Sort values so all duplicates in long and lat are dropped, dropping the NaN's in preference
    first_df_sorted = first_df.sort_values(['long', 'lat']).drop_duplicates(['long', 'lat'], keep='last')
    # Sort the new no duplicate dataframe by index
    first_df_sorted_index = first_df_sorted.sort_index()
    # Replace all the NaN back to 0
    first_df_sorted_index.replace(np.nan, '0', inplace=True)

    # ---------- Attempt number 2 to remove duplicates ---------- #
    # Load csv into dataframe
    second_df = pd.read_csv('cable_coordinates_all.csv')
    # Replace all 0 with NaN
    second_df.replace('0', np.nan, inplace=True)
    # Sort values so that all NaN are first
    second_df = second_df.sort_values('Cable_Type', na_position='first')
    # Sort values so all duplicates in long and lat are dropped, dropping the NaN's in preference
    second_df_sorted = second_df.sort_values(['long', 'lat', 'Cable_Type']).drop_duplicates(['long', 'lat'], keep='first')
    # Sort the new no duplicate dataframe by index
    second_df_sorted_index = second_df_sorted.sort_index()
    # Replace all the NaN back to 0
    second_df_sorted_index.replace(np.nan, '0', inplace=True)

    # ---------- Check which dataframe is correct ---------- #
    """
        Iterate through each row in the dataframes, comparing the current row with the previous row, checking if they 
        both contained "section", if so this dataframe is incorrect as it has not sorted the coordinates back 
        into section order as the section delineations are next to each other without coordinates inbetween them.
    """
    # Initialise variable's
    prev_row = ""
    first_df_check = False
    second_df_check = False

    # --- Check first dataframe --- #

    for index, row in first_df_sorted_index.iterrows():
        if row is str:
            # Check current row with previous row
            if prev_row == row['long'][:7] == 'section':
                # Mark first_df as being incorrect
                first_df_check = True
                break
            # Save previous row
            prev_row = row['long'][:7]

    # Re-initialise prev_row variable
    prev_row = ""

    # --- Check second dataframe --- #

    for index, row in second_df_sorted_index.iterrows():
        if row is str:
            # Check current row with previous row
            if prev_row == row['long'][:7] == 'section':
                # Mark second_df as being incorrect
                second_df_check = True
                break
            # Save previous row
            prev_row = row['long'][:7]

    # --- Convert correct dataframe into coordinates.csv --- #

    if first_df_check == False and second_df_check == True:
        first_df_sorted_index.to_csv('coordinates.csv', index=False)
    elif second_df_check == False and first_df_check == True:
        second_df_sorted_index.to_csv('coordinates.csv', index=False)
    elif first_df_check == False and second_df_check == False:
        first_df_sorted_index.to_csv('coordinates.csv', index=False)

    CSVToArray()


def CSVToArray():
    """
    This function converts a CSV into an Array, specifically unsorted_coordinates
    """
    global unsorted_coordinates

    i = 0
    # Clear unsorted_coordinates array
    unsorted_coordinates.clear()
    # Open coordinates.csv
    with open('coordinates.csv', 'r') as csvfile:
        # Iterate through each row in the csv file
        for row in csv.reader(csvfile):
            # If row is equal to the heading, skip
            if row == ["long", "lat", "Cable_Type", "Slack"]:
                continue
            # If the first 7 characters of column 0, long, does not equal section
            if row[0][:7] != "section":

                try:
                    # Attempt to append coordinates into array using an index, indicating the same section.
                    unsorted_coordinates[i].append(row)
                except IndexError:
                    # If the section, index, has not been created yet, just add the coordinates, this creates the index, section.
                    unsorted_coordinates.append([row])
            else:
                # Increment i
                i += 1


def CoordsToCSV(allCoords, cable_type, slack):
    """
    This function converts coordinates array into a csv.
    Params: allCoords: Coordinates array, cable_type: Cable type for those coordinates, slack: Slack for those coordinates
    Returns: Creates cable_coordinates_all.csv
    """
    global unsorted_coordinates
    global number
    global header_check_cable_all

    # Save coordinates from kml file into csv's
    for coords in allCoords:
        # Save section
        section = process_coordinate_string(coords.string, False, cable_type, slack)
        with open(r'cable_coordinates_all.csv', 'a', newline='') as file:
            write = csv.writer(file)
            # Write header for CSV file if a header has not already been created
            if header_check_cable_all is False:
                write.writerow(["long", "lat", "Cable_Type", "Slack"])
                header_check_cable_all = True
            else:
                # Create a delineation between each section which is unique so pandas dataframe,
                # does not automatically remove it when putting it into a dataframe later on
                write.writerow(["section" + str(number), "section" + str(number + 1), " ", " "])
                number += 1
            # Write section into CSV
            write.writerows(section)
            unsorted_coordinates.append(section)


def appendColumnToArray(coordinates, column):
    full = np.full((len(coordinates), 1), column)
    arr = np.array(coordinates)
    return np.hstack((arr, full))


def order_section(lat, lon):
    """
    This function orders a section.
    Params: lat: Latitude passed in, lon: Longitude passed in
    Returns: A coordinate or coordinates sorted
    """
    global sorted_coordinates
    global unsorted_coordinates
    cable_type = ""
    slack = ""
    # ---- Get Cable Type and Slack ---- #
    # Loop through each section unsorted_coordinates
    for item in unsorted_coordinates:
        # Loop through each item in each section in unsorted_coordinates
        for sub_item in item:
            # If lat is in sub_item, save cable_type
            if lat in sub_item:
                cable_type = sub_item[2]
                slack = sub_item[3]
                break
        # If cable_type has not been saved, loop through next section.
        if cable_type != "" and slack != "":
            break

    # Combine lat, lon, cable_type and slack together for the starting coordinates that the user selected
    starting = [lat, lon, cable_type, slack]
    sorted = coordinates_found = coordinates_added = False
    section_start = len(sorted_coordinates)
    starting_coordinates = starting.copy()
    while sorted == False:
        coordinates_found = False
        # Loop through each section in unsorted_coordinates
        for section in unsorted_coordinates:
            # If start of section is starting coordinates that user selected
            if section[0][0] == starting_coordinates[0] and section[0][1] == starting_coordinates[1]:
                # Set starting coordinates as the end of the section
                starting_coordinates = (section[(len(section) - 1)]).copy()
                if section not in sorted_coordinates:
                    # Check if only one coordinate in section
                    if len(section) == 1:
                        # Append the entire section if only one coordinate
                        sorted_coordinates.append(section)
                    else:
                        # Append section after the first coordinate in the section
                        sorted_coordinates.append(section[1:])
                coordinates_found, coordinates_added = True, True
                # Remove the section just sorted
                unsorted_coordinates.remove(section)
            # If end of section is starting coordinates that user selected
            elif section[(len(section) - 1)][0] == starting_coordinates[0] and section[(len(section) - 1)][1] == \
                    starting_coordinates[1]:
                # This reverses the order of coordinates inside the section
                section.reverse()
                # Set starting coordinates as the first coordinates, as the section has now been reversed
                starting_coordinates = (section[(len(section) - 1)]).copy()
                if len(section) <= 1:
                    sorted_coordinates.append(section)
                else:
                    sorted_coordinates.append(section[1:])
                coordinates_found, coordinates_added = True, True
                unsorted_coordinates.remove(section)

        if coordinates_found == False:
            if coordinates_added == True:
                # Check if starting coordinates are not already in sorted_coordinates
                if starting != sorted_coordinates[section_start][0]:
                    sorted_coordinates[section_start].insert(0, starting)
                    section_start = len(sorted_coordinates)
            # Set headings
            headings = ["Latitude", "Longitude","Cable_Type", "Slack", "Colour"]
            # Put sorted coordinates into sorted.csv
            with open("sorted.csv", "w+") as my_csv:
                csvWriter = csv.writer(my_csv, delimiter=',')
                csvWriter.writerow(headings)
                # Add each section to sorted.csv
                for section in sorted_coordinates:
                    csvWriter.writerows(appendColumnToArray(section, "Sorted"))
            # Put unsorted coordinates into unsorted.csv
            with open("unsorted.csv", "w+") as my_csv:
                csvWriter = csv.writer(my_csv, delimiter=',')
                csvWriter.writerow(headings)
                # Add each section to unsorted.csv
                for section in unsorted_coordinates:
                    csvWriter.writerows(appendColumnToArray(section, "Unsorted"))

            # use sorted and unsorted csv instead
            if len(unsorted_coordinates) == 0:
                sorted = True
                if os.path.exists("unsorted.csv"):
                    os.remove("unsorted.csv")
                # Return coordinates sorted
                return "Coordinates sorted, ", "RPL_" + graph_name + ".xls created"
            else:
                # Return coordinates
                return [starting_coordinates[0], starting_coordinates[1]]


def findNextPoint(lat, lon, fororback):
    """
    This function finds the next coordinate to suggest for the user or use in the auto create button.
    Parms: lat: Latitude passed in, lon: Longitude passed in, fororback: Whether to check for cable end, or go forward
    one coordinate or back one coordinate
    Returns: Next coordinate to use or coordinates sorted or not an end of cable
    """
    x = np.array([[lat, lon]])
    # If x equals Coordinates sorted then the coordinates are already sorted and do not need to find the next point.
    if x[0][0] == "Coordinates sorted, ":
        return x
    # Read unsorted.csv into dataframe called route
    route = pd.read_csv(f'unsorted.csv', dtype=str)
    # ---- Drop columns which we do not need ---- #
    route = route.drop(['Colour'], axis=1)
    route = route.drop(['Cable_Type'], axis=1)
    route = route.drop(['Slack'], axis=1)

    # Convert route to numpy array
    route = route.to_numpy()
    # Get index location of the supplied long and lat
    index = np.argmin(euclidean_distances(x, route))

    # Choose which check to perform
    if fororback == "StartOrEnd":
        # If index not equal to 0, start of cable, and final index, end of cable, return not end of cable.
        # Sometimes this is not applicable due to the format of the KML file, user must uncheck cable end checking
        if index != 0 and index+1 != len(route):
            return "Not an end of cable"
    if fororback == "forward":
        # Return coordinate on index after
        try:
            return route[index+1]
        except IndexError:
            return route[index]
    elif fororback == "back":
        # Return coordinate on index before
        try:
            return route[index-1]
        except IndexError:
            return route[index]
    else:
        # Return coordinate
        return route[index]
